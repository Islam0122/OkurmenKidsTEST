from __future__ import annotations

import io
import statistics
from typing import TYPE_CHECKING

from django.db.models import (
    Avg,
    Count,
    DurationField,
    ExpressionWrapper,
    F,
    FloatField,
    Max,
    Min,
    Q,
)
from django.db.models.functions import TruncDate

if TYPE_CHECKING:
    from .services import MultiSessionFilters

PASS_THRESHOLD = 75.0



def _base_attempt_qs(filters: "MultiSessionFilters"):
    from apps.testing.models import StudentAttempt, AttemptStatus

    qs = (
        StudentAttempt.objects
        .filter(session_id__in=filters.session_ids)
        .select_related("session__test")
    )

    # ── Optional filters ──────────────────────────────────────────────────
    if filters.date_from:
        qs = qs.filter(started_at__date__gte=filters.date_from)
    if filters.date_to:
        qs = qs.filter(started_at__date__lte=filters.date_to)
    if filters.test_id:
        qs = qs.filter(session__test_id=filters.test_id)
    if filters.session_type:
        qs = qs.filter(session__session_type=filters.session_type)
    if filters.status:
        qs = qs.filter(status=filters.status)
    if filters.min_score is not None:
        qs = qs.filter(score__gte=filters.min_score)
    if filters.max_score is not None:
        qs = qs.filter(score__lte=filters.max_score)

    return qs


def get_multi_session_kpis(filters: "MultiSessionFilters") -> dict:
    from apps.testing.models import (
        TestSession, StudentAttempt, AttemptStatus,
    )

    sessions = list(
        TestSession.objects
        .filter(pk__in=filters.session_ids)
        .select_related("test")
        .values("id", "title", "key", "session_type", "test__title", "created_at", "expires_at")
    )
    if not sessions:
        return {}

    session_titles = [s["title"] or s["key"][:16] for s in sessions]
    test_titles    = sorted({s["test__title"] for s in sessions})

    base_qs = _base_attempt_qs(filters)

    agg = base_qs.aggregate(
        total_count   = Count("id"),
        active_count  = Count("id", filter=Q(status=AttemptStatus.ACTIVE)),
        finished_count= Count("id", filter=Q(status=AttemptStatus.FINISHED)),
        expired_count = Count("id", filter=Q(status=AttemptStatus.EXPIRED)),
        passed_count  = Count("id", filter=Q(status=AttemptStatus.FINISHED, score__gte=PASS_THRESHOLD)),
        failed_count  = Count("id", filter=Q(status=AttemptStatus.FINISHED, score__lt=PASS_THRESHOLD)),
        avg_score     = Avg("score",  filter=Q(status=AttemptStatus.FINISHED)),
        max_score     = Max("score",  filter=Q(status=AttemptStatus.FINISHED)),
        min_score     = Min("score",  filter=Q(status=AttemptStatus.FINISHED)),
        avg_duration  = Avg(
            ExpressionWrapper(
                F("finished_at") - F("started_at"),
                output_field=DurationField(),
            ),
            filter=Q(status=AttemptStatus.FINISHED, finished_at__isnull=False),
        ),
    )

    unique_students = (
        base_qs
        .values("student_name")
        .distinct()
        .count()
    )

    finished_scores = list(
        base_qs
        .filter(status=AttemptStatus.FINISHED)
        .values_list("score", flat=True)
        .order_by("score")
    )
    median_score = (
        round(statistics.median(finished_scores), 1)
        if finished_scores else 0.0
    )

    finished   = agg["finished_count"] or 0
    total      = agg["total_count"]    or 0
    pass_rate  = round(finished / total * 100, 1) if total > 0 else 0.0

    avg_dur_secs = None
    if agg["avg_duration"]:
        avg_dur_secs = int(agg["avg_duration"].total_seconds())

    per_session = list(
        base_qs
        .filter(status=AttemptStatus.FINISHED)
        .values("session_id", "session__title", "session__key")
        .annotate(s_avg=Avg("score"))
        .order_by("-s_avg")
    )

    best_session  = per_session[0]  if per_session else None
    worst_session = per_session[-1] if per_session else None

    return {
        "session_count":       len(sessions),
        "session_titles":      session_titles,
        "test_titles":         test_titles,
        "total_count":         total,
        "active_count":        agg["active_count"]   or 0,
        "finished_count":      finished,
        "expired_count":       agg["expired_count"]  or 0,
        "unique_students":     unique_students,
        "avg_score":           round(agg["avg_score"] or 0, 1),
        "median_score":        median_score,
        "max_score":           round(agg["max_score"] or 0, 1),
        "min_score":           round(agg["min_score"] or 0, 1),
        "passed_count":        agg["passed_count"]   or 0,
        "failed_count":        agg["failed_count"]   or 0,
        "pass_threshold":      PASS_THRESHOLD,
        "pass_rate":           pass_rate,
        "fail_rate":           round(100 - pass_rate, 1),
        "avg_duration_seconds": avg_dur_secs,
        "best_session":        _fmt_session_label(best_session),
        "best_session_avg":    round(best_session["s_avg"] or 0, 1) if best_session else 0,
        "worst_session":       _fmt_session_label(worst_session),
        "worst_session_avg":   round(worst_session["s_avg"] or 0, 1) if worst_session else 0,
        "per_session":         per_session,
    }


def _fmt_session_label(row: dict | None) -> str:
    if row is None:
        return "—"
    return row.get("session__title") or (row.get("session__key", "")[:16] + "…")



def get_multi_session_ranking(filters: "MultiSessionFilters") -> list[dict]:
    from apps.testing.models import AttemptStatus

    qs = (
        _base_attempt_qs(filters)
        .filter(status=AttemptStatus.FINISHED)
        .annotate(
            duration=ExpressionWrapper(
                F("finished_at") - F("started_at"),
                output_field=DurationField(),
            )
        )
        .order_by("-score", "duration")
        .values(
            "id", "student_name", "score", "duration",
            "started_at", "finished_at",
            "session_id",
            "session__title", "session__key",
            "session__test__title",
        )
    )

    rows = list(qs)

    if filters.dedup_mode == "best":
        rows = _dedup_best(rows)

    result = []
    for rank, row in enumerate(rows, start=1):
        result.append({
            "rank":             rank,
            "attempt_id":       row["id"],
            "student_name":     row["student_name"],
            "score":            row["score"],
            "duration_seconds": _td_to_secs(row["duration"]),
            "started_at":       row["started_at"],
            "finished_at":      row["finished_at"],
            "session_title":    row["session__title"] or row["session__key"][:16],
            "test_title":       row["session__test__title"],
        })
    return result


def _dedup_best(rows: list[dict]) -> list[dict]:
    seen: dict[str, dict] = {}
    for row in rows:
        name = row["student_name"]
        if name not in seen:
            seen[name] = row
    return sorted(seen.values(), key=lambda r: (-r["score"], _td_to_secs(r["duration"]) or 0))


def _td_to_secs(td) -> float | None:
    if td is None:
        return None
    return round(td.total_seconds(), 1)



def get_multi_session_question_breakdown(filters: "MultiSessionFilters") -> list[dict]:

    from apps.testing.models import Answer

    rows = (
        Answer.objects
        .filter(attempt__session_id__in=filters.session_ids)
        .values(
            "question_id",
            "question__text",
            "question__question_type",
            "question__difficulty",
            "question__order",
        )
        .annotate(
            total_answers    = Count("id"),
            correct_answers  = Count("id", filter=Q(is_correct=True)),
            incorrect_answers= Count("id", filter=Q(is_correct=False)),
            pending_answers  = Count("id", filter=Q(is_correct__isnull=True)),
        )
        .order_by("question__order", "question__created_at")
    )

    if filters.date_from:
        rows = rows.filter(answered_at__date__gte=filters.date_from)
    if filters.date_to:
        rows = rows.filter(answered_at__date__lte=filters.date_to)

    result = []
    for r in rows:
        total        = r["total_answers"] or 1
        correct_rate = round(r["correct_answers"] / total * 100, 1)
        result.append({
            "question_id":     r["question_id"],
            "question_text":   r["question__text"][:80],
            "question_type":   r["question__question_type"],
            "difficulty":      r["question__difficulty"],
            "total_answers":   r["total_answers"],
            "correct_answers": r["correct_answers"],
            "incorrect_answers": r["incorrect_answers"],
            "pending_answers": r["pending_answers"],
            "correct_rate":    correct_rate,
        })
    return result



def get_score_buckets(filters: "MultiSessionFilters") -> list[dict]:
    from apps.testing.models import AttemptStatus

    scores = list(
        _base_attempt_qs(filters)
        .filter(status=AttemptStatus.FINISHED)
        .values_list("score", flat=True)
    )
    buckets = {f"{i*10}-{i*10+10}": 0 for i in range(10)}
    for sc in scores:
        idx = min(int(sc // 10), 9)
        key = f"{idx*10}-{idx*10+10}"
        buckets[key] += 1
    return [{"range": k, "count": v} for k, v in buckets.items()]


def get_score_by_session(filters: "MultiSessionFilters") -> list[dict]:
    from apps.testing.models import AttemptStatus

    rows = (
        _base_attempt_qs(filters)
        .filter(status=AttemptStatus.FINISHED)
        .values("session_id", "session__title", "session__key")
        .annotate(avg_score=Avg("score"), count=Count("id"))
        .order_by("session__created_at")
    )
    return [
        {
            "label":     r["session__title"] or r["session__key"][:12],
            "avg_score": round(r["avg_score"] or 0, 1),
            "count":     r["count"],
        }
        for r in rows
    ]


def get_attempts_by_day(filters: "MultiSessionFilters") -> list[dict]:
    rows = (
        _base_attempt_qs(filters)
        .annotate(day=TruncDate("started_at"))
        .values("day")
        .annotate(count=Count("id"))
        .order_by("day")
    )
    return [{"day": str(r["day"]), "count": r["count"]} for r in rows]



def build_multi_excel(data: dict, filters: "MultiSessionFilters") -> bytes:
    import openpyxl

    wb  = openpyxl.Workbook()
    kpis = data["kpis"]

    ws1 = wb.active
    ws1.title = "KPI сводка"
    _hdr(ws1, ["Метрика", "Значение"])
    for row in [
        ("Тестов",                  ", ".join(kpis.get("test_titles", []))),
        ("Выбрано сессий",          kpis["session_count"]),
        ("Всего попыток",           kpis["total_count"]),
        ("Завершено",               kpis["finished_count"]),
        ("Активных",                kpis["active_count"]),
        ("Просрочено",              kpis["expired_count"]),
        ("Уникальных студентов",    kpis["unique_students"]),
        ("Средний балл",            kpis["avg_score"]),
        ("Медианный балл",          kpis["median_score"]),
        ("Максимальный балл",       kpis["max_score"]),
        ("Минимальный балл",        kpis["min_score"]),
        ("Прошли (≥75%)",           kpis["passed_count"]),
        ("Не прошли (<75%)",        kpis["failed_count"]),
        ("% прохождения",           f"{kpis['pass_rate']}%"),
        ("% непрохождения",         f"{kpis['fail_rate']}%"),
        ("Лучшая сессия",           kpis.get("best_session", "—")),
        ("Лучший сред. балл",       kpis.get("best_session_avg", "—")),
        ("Худшая сессия",           kpis.get("worst_session", "—")),
        ("Худший сред. балл",       kpis.get("worst_session_avg", "—")),
        ("Среднее время",           kpis.get("avg_duration_fmt", "—")),
    ]:
        ws1.append(row)
    _autowidth(ws1)

    ws2 = wb.create_sheet("Рейтинг")
    _hdr(ws2, ["Место", "Студент", "Тест", "Сессия", "Балл", "Время", "Начато", "Завершено"])
    for row in data.get("ranking", []):
        ws2.append([
            row["rank"],
            row["student_name"],
            row.get("test_title", ""),
            row.get("session_title", ""),
            row["score"],
            row.get("duration_fmt", "—"),
            str(row["started_at"])[:19] if row.get("started_at") else "—",
            str(row["finished_at"])[:19] if row.get("finished_at") else "—",
        ])
    _autowidth(ws2)

    ws3 = wb.create_sheet("Вопросы")
    _hdr(ws3, ["Вопрос", "Тип", "Сложность", "Всего", "Верно", "Неверно", "% верных"])
    for b in data.get("breakdown", []):
        ws3.append([
            b["question_text"],
            b["question_type"],
            b["difficulty"],
            b["total_answers"],
            b["correct_answers"],
            b["incorrect_answers"],
            b["correct_rate"],
        ])
    _autowidth(ws3)

    ws4 = wb.create_sheet("По сессиям")
    _hdr(ws4, ["Сессия", "Средний балл"])
    for row in kpis.get("per_session", []):
        label = row.get("session__title") or row.get("session__key", "")[:16]
        ws4.append([label, round(row.get("s_avg") or 0, 1)])
    _autowidth(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _hdr(ws, cols: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(cols)
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)