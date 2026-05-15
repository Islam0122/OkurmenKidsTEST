
from __future__ import annotations

import json
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.admin import site as _admin_site
from django.db.models import Q          # ← single import, no re-import inside views
from django.http import HttpResponse
from django.shortcuts import render

from .selectors import (
    get_sessions_for_selector,
    get_session_kpis,
    get_session_ranking,
    get_question_breakdown,
)
from .aggregations import (
    get_multi_session_kpis,
    get_multi_session_ranking,
    get_multi_session_question_breakdown,
)


class _FakeOpts:
    """Minimal opts shim for Jazzmin breadcrumb rendering."""
    app_label           = "analytics"
    model_name          = "sessionanalytics"
    verbose_name        = "Аналитика сессий"
    verbose_name_plural = "Аналитика сессий"


def _fmt_duration(seconds: int | None) -> str:
    if seconds is None:
        return "—"
    m, s = divmod(int(seconds), 60)
    return f"{m}м {s}с" if m else f"{s}с"


@staff_member_required
def session_list_view(request):
    q = request.GET.get("q", "").strip()

    sessions_qs = get_sessions_for_selector()
    if q:
        sessions_qs = sessions_qs.filter(
            Q(title__icontains=q)
            | Q(test__title__icontains=q)
            | Q(key__icontains=q)
        )

    session_list = list(sessions_qs[:200])

    ctx = {
        **_admin_site.each_context(request),
        "title": "Аналитика сессий",
        "opts": _FakeOpts(),
        "sessions": session_list,
        "search_query": q,
    }
    return render(request, "admin/analytics/session_list.html", ctx)


@staff_member_required
def session_detail_view(request, session_id: str):

    kpis = get_session_kpis(str(session_id))
    if not kpis:
        from django.http import Http404
        raise Http404("Session not found.")

    ranking   = get_session_ranking(str(session_id))
    breakdown = get_question_breakdown(str(session_id))

    for row in ranking:
        row["duration_fmt"] = _fmt_duration(row["duration_seconds"])

    kpis["avg_duration_fmt"] = _fmt_duration(kpis.get("avg_duration_seconds"))

    score_buckets = _score_distribution(str(session_id))

    ctx = {
        **_admin_site.each_context(request),
        "title": f"Аналитика: {kpis['session_title']}",
        "opts": _FakeOpts(),
        "kpis": kpis,
        "ranking": ranking,
        "breakdown": breakdown,
        "score_buckets_json": json.dumps(score_buckets),
        "session_id": str(session_id),
    }
    return render(request, "admin/analytics/session_detail.html", ctx)


@staff_member_required
def session_export_view(request, session_id: str):
    kpis = get_session_kpis(str(session_id))
    if not kpis:
        from django.http import Http404
        raise Http404("Session not found.")

    ranking   = get_session_ranking(str(session_id))
    breakdown = get_question_breakdown(str(session_id))
    xlsx      = _build_excel(kpis, ranking, breakdown)

    label = kpis["session_title"][:30]
    resp  = HttpResponse(
        xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="analytics_{label}.xlsx"'
    return resp


@staff_member_required
def multi_session_analytics_view(request):
    from .services import MultiSessionService, MultiSessionFilters
    raw_ids    = request.GET.get("sessions", "").strip()
    session_ids = [s.strip() for s in raw_ids.split(",") if s.strip()]

    if not session_ids:
        ctx = {
            **_admin_site.each_context(request),
            "title": "Мультисессионная аналитика",
            "opts": _FakeOpts(),
            "no_sessions": True,
            "sessions": list(get_sessions_for_selector()[:200]),
        }
        return render(request, "admin/analytics/multi_analytics.html", ctx)

    filters = MultiSessionFilters(
        session_ids  = session_ids,
        date_from    = request.GET.get("date_from", "").strip() or None,
        date_to      = request.GET.get("date_to",   "").strip() or None,
        test_id      = request.GET.get("test_id",   "").strip() or None,
        session_type = request.GET.get("session_type", "").strip() or None,
        status       = request.GET.get("status",    "").strip() or None,
        min_score    = _safe_float(request.GET.get("min_score")),
        max_score    = _safe_float(request.GET.get("max_score")),
        dedup_mode   = request.GET.get("dedup", "best"),   # "best" | "all"
    )

    data = MultiSessionService.get_analytics(filters)

    for row in data["ranking"]:
        row["duration_fmt"] = _fmt_duration(row.get("duration_seconds"))

    data["kpis"]["avg_duration_fmt"] = _fmt_duration(
        data["kpis"].get("avg_duration_seconds")
    )

    ctx = {
        **_admin_site.each_context(request),
        "title": f"Мультисессионная аналитика ({len(session_ids)} сессий)",
        "opts": _FakeOpts(),
        "data": data,
        "filters": filters,
        "session_ids_csv": raw_ids,
        "score_buckets_json": json.dumps(data.get("score_buckets", [])),
        "score_by_session_json": json.dumps(data.get("score_by_session", [])),
        "attempts_by_day_json": json.dumps(data.get("attempts_by_day", [])),
    }
    return render(request, "admin/analytics/multi_analytics.html", ctx)


@staff_member_required
def multi_session_export_view(request):
    from .services import MultiSessionService, MultiSessionFilters
    from .aggregations import build_multi_excel

    raw_ids     = request.GET.get("sessions", "").strip()
    session_ids = [s.strip() for s in raw_ids.split(",") if s.strip()]

    if not session_ids:
        from django.http import Http404
        raise Http404("No sessions specified.")

    filters = MultiSessionFilters(
        session_ids  = session_ids,
        date_from    = request.GET.get("date_from", "").strip() or None,
        date_to      = request.GET.get("date_to",   "").strip() or None,
        test_id      = request.GET.get("test_id",   "").strip() or None,
        session_type = request.GET.get("session_type", "").strip() or None,
        status       = request.GET.get("status",    "").strip() or None,
        min_score    = _safe_float(request.GET.get("min_score")),
        max_score    = _safe_float(request.GET.get("max_score")),
        dedup_mode   = request.GET.get("dedup", "best"),
    )

    data = MultiSessionService.get_analytics(filters)
    xlsx = build_multi_excel(data, filters)

    resp = HttpResponse(
        xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="multi_analytics_{len(session_ids)}_sessions.xlsx"'
    )
    return resp


def _safe_float(val) -> float | None:
    try:
        return float(val) if val not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _score_distribution(session_id: str) -> list[dict]:
    from apps.testing.models import StudentAttempt, AttemptStatus

    scores = list(
        StudentAttempt.objects
        .filter(session_id=session_id, status=AttemptStatus.FINISHED)
        .values_list("score", flat=True)
    )

    buckets = {f"{i*10}-{i*10+10}": 0 for i in range(10)}
    for sc in scores:
        idx = min(int(sc // 10), 9)
        key = f"{idx*10}-{idx*10+10}"
        buckets[key] += 1

    return [{"range": k, "count": v} for k, v in buckets.items()]


def _build_excel(kpis: dict, ranking: list, breakdown: list) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "KPI Summary"

    _hdr(ws1, ["Метрика", "Значение"])
    for row in [
        ("Тест",                   kpis["test_title"]),
        ("Сессия",                 kpis["session_title"]),
        ("Тип сессии",             kpis["session_type_display"]),
        ("Всего вопросов",         kpis["total_questions"]),
        ("Всего попыток",          kpis["total_count"]),
        ("Активных попыток",       kpis["active_count"]),
        ("Завершённых попыток",    kpis["finished_count"]),
        ("Просроченных попыток",   kpis["expired_count"]),
        ("Средний балл",           kpis["avg_score"]),
        ("Максимальный балл",      kpis["max_score"]),
        ("Минимальный балл",       kpis["min_score"]),
        ("Прошли (≥75%)",          kpis["passed_count"]),
        ("Не прошли (<75%)",       kpis["failed_count"]),
        ("Процент прохождения",    f"{kpis['completion_rate']}%"),
        ("Среднее время",          kpis.get("avg_duration_fmt", "—")),
    ]:
        ws1.append(row)
    _autowidth(ws1)

    ws2 = wb.create_sheet("Рейтинг")
    _hdr(ws2, ["Место", "Студент", "Балл", "Время", "Начато", "Завершено"])
    for row in ranking:
        ws2.append([
            row["rank"],
            row["student_name"],
            row["score"],
            row.get("duration_fmt", "—"),
            str(row["started_at"])[:19] if row["started_at"] else "—",
            str(row["finished_at"])[:19] if row["finished_at"] else "—",
        ])
    _autowidth(ws2)

    ws3 = wb.create_sheet("Вопросы")
    _hdr(ws3, ["Вопрос", "Тип", "Сложность", "Всего", "Верно", "Неверно", "% верных"])
    for b in breakdown:
        ws3.append([
            b["question_text"],
            b["question_type"],
            b["difficulty"],
            b["total_answers"],
            b["correct_answers"],
            b["incorrect_answers"],
            b["correct_rate"],
        ])
    _autowidth(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _hdr(ws, cols):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(cols)
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)