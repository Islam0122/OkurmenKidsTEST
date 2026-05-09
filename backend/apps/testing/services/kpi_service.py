"""
services/kpi_service.py — KPI Dashboard data layer for OkurmenKids.

All heavy lifting happens here: ORM aggregations, caching, export.
admin.py calls only KPIService.get_dashboard(filters).
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date, timedelta
from typing import Any

from django.core.cache import cache
from django.db.models import (
    Avg, Case, Count, DurationField, ExpressionWrapper, F, FloatField,
    IntegerField, Max, Min, Q, Sum, Value, When,
)
from django.db.models.functions import TruncDate
from django.utils import timezone

logger = logging.getLogger(__name__)

# ── Cache TTL ─────────────────────────────────────────────────────────────────
CACHE_TTL = 60 * 5  # 5 minutes


# ── Date range helpers ────────────────────────────────────────────────────────

def _date_range(
    period: str,
    date_from: str | None = None,
    date_to: str | None = None,
) -> tuple[Any, Any]:
    """
    Returns (start, end) as timezone-aware datetimes.
    Priority: custom date_from/date_to > period preset > None/None (all time).
    """
    from datetime import datetime

    now = timezone.now()

    # Custom range has highest priority
    if date_from or date_to:
        start = end = None
        if date_from:
            try:
                d = datetime.strptime(date_from, "%Y-%m-%d")
                start = timezone.make_aware(d.replace(hour=0, minute=0, second=0))
            except ValueError:
                pass
        if date_to:
            try:
                d = datetime.strptime(date_to, "%Y-%m-%d")
                end = timezone.make_aware(d.replace(hour=23, minute=59, second=59))
            except ValueError:
                pass
        return start, end

    # Preset periods
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now
    if period == "7d":
        return now - timedelta(days=7), now
    if period == "30d":
        return now - timedelta(days=30), now
    return None, None  # "all"


# ── Filters dataclass ────────────────────────────────────────────────────────

@dataclass
class KPIFilters:
    period:    str = "all"        # today | 7d | 30d | all | custom
    test_id:   str | None = None  # UUID string or None
    date_from: str | None = None  # "YYYY-MM-DD"
    date_to:   str | None = None  # "YYYY-MM-DD"

    @property
    def is_custom_range(self) -> bool:
        return bool(self.date_from or self.date_to)

    @property
    def cache_key(self) -> str:
        return (
            f"kpi_dashboard:{self.period}:{self.test_id or 'all'}"
            f":{self.date_from or ''}:{self.date_to or ''}"
        )

    def _range(self) -> tuple[Any, Any]:
        return _date_range(self.period, self.date_from, self.date_to)

    def attempt_q(self) -> Q:
        """Base Q filter for StudentAttempt queryset."""
        q = Q()
        start, end = self._range()
        if start:
            q &= Q(started_at__gte=start)
        if end:
            q &= Q(started_at__lte=end)
        if self.test_id:
            q &= Q(session__test_id=self.test_id)
        return q

    def session_q(self) -> Q:
        q = Q()
        start, end = self._range()
        if start:
            q &= Q(created_at__gte=start)
        if end:
            q &= Q(created_at__lte=end)
        if self.test_id:
            q &= Q(test_id=self.test_id)
        return q

    def answer_q(self) -> Q:
        q = Q()
        start, end = self._range()
        if start:
            q &= Q(answered_at__gte=start)
        if end:
            q &= Q(answered_at__lte=end)
        if self.test_id:
            q &= Q(attempt__session__test_id=self.test_id)
        return q


# ── Main service ──────────────────────────────────────────────────────────────

class KPIService:

    @staticmethod
    def get_dashboard(filters: KPIFilters) -> dict:
        cached = cache.get(filters.cache_key)
        if cached is not None:
            return cached

        data = {
            "summary":         KPIService._summary(filters),
            "session_types":   KPIService._by_session_type(filters),
            "difficulties":    KPIService._by_difficulty(filters),
            "question_types":  KPIService._by_question_type(filters),
            "languages":       KPIService._by_language(filters),
            "top_students":    KPIService._top_students(filters),
            "active_students": KPIService._most_active_students(filters),
            "tests_kpi":       KPIService._tests_kpi(filters),
            "ai_kpi":          KPIService._ai_kpi(filters),
            "chart_attempts":  KPIService._chart_attempts_per_day(filters),
            "chart_scores":    KPIService._chart_avg_score_per_day(filters),
            "all_tests":       KPIService._all_tests(),
        }

        cache.set(filters.cache_key, data, CACHE_TTL)
        return data

    # ── 1. Summary cards ──────────────────────────────────────────────────────

    @staticmethod
    def _summary(f: KPIFilters) -> dict:
        from apps.testing.models import (
            Test, Question, TestSession, StudentAttempt, Answer,
            AttemptStatus, GradingStatus,
        )

        now = timezone.now()

        # --- Tests & Questions (not date-filtered — structural data)
        test_qs = Test.objects.all()
        if f.test_id:
            test_qs = test_qs.filter(pk=f.test_id)

        total_tests     = test_qs.count()
        total_questions = Question.objects.filter(
            test__in=test_qs
        ).count()

        # --- Sessions
        session_qs  = TestSession.objects.filter(f.session_q())
        total_sess  = session_qs.count()
        active_sess = session_qs.filter(is_active=True, expires_at__gt=now).count()
        done_sess   = session_qs.filter(
            Q(is_active=False) | Q(expires_at__lte=now)
        ).count()

        # --- Attempts
        attempt_qs      = StudentAttempt.objects.filter(f.attempt_q())
        total_attempts  = attempt_qs.count()
        done_attempts   = attempt_qs.filter(status=AttemptStatus.FINISHED).count()
        exp_attempts    = attempt_qs.filter(status=AttemptStatus.EXPIRED).count()

        # Unique students (distinct student_name across filtered attempts)
        total_students = attempt_qs.values("student_name").distinct().count()

        # Score stats (finished only)
        score_agg = attempt_qs.filter(
            status=AttemptStatus.FINISHED
        ).aggregate(
            avg=Avg("score"),
            mx=Max("score"),
            mn=Min("score"),
        )

        # Duration average (seconds)
        dur_agg = attempt_qs.filter(
            status=AttemptStatus.FINISHED,
            finished_at__isnull=False,
        ).annotate(
            dur=ExpressionWrapper(
                F("finished_at") - F("started_at"),
                output_field=DurationField(),
            )
        ).aggregate(avg_dur=Avg("dur"))

        avg_dur_secs = None
        if dur_agg["avg_dur"]:
            avg_dur_secs = int(dur_agg["avg_dur"].total_seconds())

        # --- Grading breakdown
        answer_qs    = Answer.objects.filter(f.answer_q())
        ai_count     = answer_qs.filter(grading_status=GradingStatus.AI).count()
        done_count   = answer_qs.filter(grading_status=GradingStatus.DONE).count()
        auto_count   = answer_qs.filter(grading_status=GradingStatus.AUTO).count()
        manual_count = answer_qs.filter(grading_status=GradingStatus.MANUAL).count()
        failed_count = answer_qs.filter(grading_status=GradingStatus.FAILED).count()

        return {
            "total_tests":      total_tests,
            "total_questions":  total_questions,
            "total_sessions":   total_sess,
            "active_sessions":  active_sess,
            "done_sessions":    done_sess,
            "total_students":   total_students,
            "total_attempts":   total_attempts,
            "done_attempts":    done_attempts,
            "expired_attempts": exp_attempts,
            "avg_score":        round(score_agg["avg"] or 0, 1),
            "max_score":        round(score_agg["mx"] or 0, 1),
            "min_score":        round(score_agg["mn"] or 0, 1),
            "avg_duration_secs": avg_dur_secs,
            "ai_checks":        ai_count + done_count,
            "auto_checks":      auto_count,
            "manual_checks":    manual_count,
            "failed_ai":        failed_count,
        }

    # ── 2. By session type ────────────────────────────────────────────────────

    @staticmethod
    def _by_session_type(f: KPIFilters) -> list[dict]:
        from apps.testing.models import StudentAttempt, AttemptStatus

        rows = (
            StudentAttempt.objects
            .filter(f.attempt_q())
            .values("session__session_type")
            .annotate(
                sessions=Count("session_id", distinct=True),
                students=Count("student_name", distinct=True),
                attempts=Count("id"),
                avg_score=Avg("score", filter=Q(status=AttemptStatus.FINISHED)),
                finished=Count("id", filter=Q(status=AttemptStatus.FINISHED)),
                total=Count("id"),
            )
        )

        result = []
        for r in rows:
            total    = r["total"] or 1
            finished = r["finished"] or 0
            result.append({
                "session_type": r["session__session_type"],
                "sessions":     r["sessions"],
                "students":     r["students"],
                "attempts":     r["attempts"],
                "avg_score":    round(r["avg_score"] or 0, 1),
                "pass_rate":    round(finished / total * 100, 1),
            })
        return result

    # ── 3. By difficulty ──────────────────────────────────────────────────────

    @staticmethod
    def _by_difficulty(f: KPIFilters) -> list[dict]:
        from apps.testing.models import Question

        qs = Question.objects.all()
        if f.test_id:
            qs = qs.filter(test_id=f.test_id)

        total = qs.count() or 1
        rows  = (
            qs.values("difficulty")
              .annotate(count=Count("id"))
              .order_by("difficulty")
        )
        return [
            {
                "difficulty": r["difficulty"],
                "count":      r["count"],
                "percent":    round(r["count"] / total * 100, 1),
            }
            for r in rows
        ]

    # ── 4. By question type ───────────────────────────────────────────────────

    @staticmethod
    def _by_question_type(f: KPIFilters) -> list[dict]:
        from apps.testing.models import Question

        qs = Question.objects.all()
        if f.test_id:
            qs = qs.filter(test_id=f.test_id)

        total = qs.count() or 1
        rows  = (
            qs.values("question_type")
              .annotate(count=Count("id"))
              .order_by("-count")
        )
        return [
            {
                "question_type": r["question_type"],
                "count":         r["count"],
                "percent":       round(r["count"] / total * 100, 1),
            }
            for r in rows
        ]

    # ── 5. By language ────────────────────────────────────────────────────────

    @staticmethod
    def _by_language(f: KPIFilters) -> list[dict]:
        from apps.testing.models import Question

        qs = Question.objects.filter(question_type="code").exclude(language="")
        if f.test_id:
            qs = qs.filter(test_id=f.test_id)

        rows = (
            qs.values("language")
              .annotate(count=Count("id"))
              .order_by("-count")
        )
        return [{"language": r["language"], "count": r["count"]} for r in rows]

    # ── 6. Top students (by score) ────────────────────────────────────────────

    @staticmethod
    def _top_students(f: KPIFilters, limit: int = 10) -> list[dict]:
        from apps.testing.models import StudentAttempt, AttemptStatus

        rows = (
            StudentAttempt.objects
            .filter(f.attempt_q(), status=AttemptStatus.FINISHED)
            .values("student_name")
            .annotate(
                attempts=Count("id"),
                avg_score=Avg("score"),
                best_score=Max("score"),
                last_activity=Max("finished_at"),
                avg_dur=Avg(
                    ExpressionWrapper(
                        F("finished_at") - F("started_at"),
                        output_field=DurationField(),
                    )
                ),
            )
            .order_by("-avg_score")[:limit]
        )
        return [
            {
                "student_name":    r["student_name"],
                "attempts":        r["attempts"],
                "avg_score":       round(r["avg_score"] or 0, 1),
                "best_score":      round(r["best_score"] or 0, 1),
                "last_activity":   r["last_activity"],
                "avg_dur_secs":    int(r["avg_dur"].total_seconds()) if r["avg_dur"] else None,
            }
            for r in rows
        ]

    # ── 7. Most active students ───────────────────────────────────────────────

    @staticmethod
    def _most_active_students(f: KPIFilters, limit: int = 10) -> list[dict]:
        from apps.testing.models import StudentAttempt, AttemptStatus

        rows = (
            StudentAttempt.objects
            .filter(f.attempt_q())
            .values("student_name")
            .annotate(
                attempts=Count("id"),
                avg_score=Avg("score", filter=Q(status=AttemptStatus.FINISHED)),
                best_score=Max("score"),
                last_activity=Max("started_at"),
            )
            .order_by("-attempts")[:limit]
        )
        return [
            {
                "student_name":  r["student_name"],
                "attempts":      r["attempts"],
                "avg_score":     round(r["avg_score"] or 0, 1),
                "best_score":    round(r["best_score"] or 0, 1),
                "last_activity": r["last_activity"],
            }
            for r in rows
        ]

    # ── 8. Tests KPI ──────────────────────────────────────────────────────────

    @staticmethod
    def _tests_kpi(f: KPIFilters) -> list[dict]:
        from apps.testing.models import Test, StudentAttempt, AttemptStatus

        PASS = 75.0

        test_qs = Test.objects.all()
        if f.test_id:
            test_qs = test_qs.filter(pk=f.test_id)

        result = []
        for test in test_qs.prefetch_related("questions"):
            att_qs = StudentAttempt.objects.filter(
                f.attempt_q(), session__test=test
            )
            agg = att_qs.aggregate(
                sessions=Count("session_id", distinct=True),
                total=Count("id"),
                finished=Count("id", filter=Q(status=AttemptStatus.FINISHED)),
                passed=Count("id", filter=Q(status=AttemptStatus.FINISHED, score__gte=PASS)),
                failed=Count("id", filter=Q(status=AttemptStatus.FINISHED, score__lt=PASS)),
                avg_score=Avg("score", filter=Q(status=AttemptStatus.FINISHED)),
                best_score=Max("score"),
            )
            total    = agg["total"] or 1
            finished = agg["finished"] or 0
            result.append({
                "test_id":       str(test.id),
                "test_title":    test.title,
                "questions":     test.questions.count(),
                "sessions":      agg["sessions"] or 0,
                "total_attempts": agg["total"] or 0,
                "avg_score":     round(agg["avg_score"] or 0, 1),
                "best_score":    round(agg["best_score"] or 0, 1),
                "pass_rate":     round(finished / total * 100, 1),
                "passed":        agg["passed"] or 0,
                "failed":        agg["failed"] or 0,
            })

        return sorted(result, key=lambda x: x["total_attempts"], reverse=True)

    # ── 9. AI grading KPI ─────────────────────────────────────────────────────

    @staticmethod
    def _ai_kpi(f: KPIFilters) -> dict:
        from apps.testing.models import Answer, GradingStatus

        qs = Answer.objects.filter(f.answer_q())
        agg = qs.aggregate(
            pending=Count("id", filter=Q(grading_status=GradingStatus.PENDING)),
            processing=Count("id", filter=Q(grading_status=GradingStatus.PROCESSING)),
            done=Count("id", filter=Q(grading_status=GradingStatus.DONE)),
            failed=Count("id", filter=Q(grading_status=GradingStatus.FAILED)),
            avg_ai_score=Avg("ai_score", filter=Q(ai_score__isnull=False)),
            avg_confidence=Avg("ai_confidence", filter=Q(ai_confidence__isnull=False)),
        )
        return {
            "pending":         agg["pending"] or 0,
            "processing":      agg["processing"] or 0,
            "done":            agg["done"] or 0,
            "failed":          agg["failed"] or 0,
            "avg_ai_score":    round(agg["avg_ai_score"] or 0, 2),
            "avg_confidence":  round(agg["avg_confidence"] or 0, 2),
        }

    # ── 10. Chart: attempts per day ───────────────────────────────────────────

    @staticmethod
    def _chart_attempts_per_day(f: KPIFilters) -> list[dict]:
        from apps.testing.models import StudentAttempt

        rows = (
            StudentAttempt.objects
            .filter(f.attempt_q())
            .annotate(day=TruncDate("started_at"))
            .values("day")
            .annotate(count=Count("id"))
            .order_by("day")
        )
        return [{"day": str(r["day"]), "count": r["count"]} for r in rows]

    # ── 11. Chart: avg score per day ──────────────────────────────────────────

    @staticmethod
    def _chart_avg_score_per_day(f: KPIFilters) -> list[dict]:
        from apps.testing.models import StudentAttempt, AttemptStatus

        rows = (
            StudentAttempt.objects
            .filter(f.attempt_q(), status=AttemptStatus.FINISHED)
            .annotate(day=TruncDate("started_at"))
            .values("day")
            .annotate(avg=Avg("score"))
            .order_by("day")
        )
        return [{"day": str(r["day"]), "avg": round(r["avg"] or 0, 1)} for r in rows]

    # ── Helpers ───────────────────────────────────────────────────────────────

    @staticmethod
    def _all_tests() -> list[dict]:
        from apps.testing.models import Test
        return list(Test.objects.values("id", "title").order_by("title"))

    @staticmethod
    def invalidate(filters: KPIFilters | None = None) -> None:
        """Call after any write operation to flush relevant cache keys."""
        periods = ["all", "today", "7d", "30d"]
        test_ids = [None]
        if filters and filters.test_id:
            test_ids.append(filters.test_id)
        for period in periods:
            for tid in test_ids:
                key = f"kpi_dashboard:{period}:{tid or 'all'}::"
                cache.delete(key)
        # Also bust the specific custom-range key if provided
        if filters and filters.is_custom_range:
            cache.delete(filters.cache_key)


# ── Excel / CSV export ────────────────────────────────────────────────────────

def export_kpi_excel(data: dict) -> bytes:
    """Export KPI summary to .xlsx. Returns bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка KPI"
    _header(ws, ["Метрика", "Значение"])
    s = data["summary"]
    rows = [
        ("Всего тестов",            s["total_tests"]),
        ("Всего вопросов",          s["total_questions"]),
        ("Всего сессий",            s["total_sessions"]),
        ("Активных сессий",         s["active_sessions"]),
        ("Завершённых сессий",      s["done_sessions"]),
        ("Уникальных студентов",    s["total_students"]),
        ("Всего попыток",           s["total_attempts"]),
        ("Завершённых попыток",     s["done_attempts"]),
        ("Просроченных попыток",    s["expired_attempts"]),
        ("Средний балл",            s["avg_score"]),
        ("Максимальный балл",       s["max_score"]),
        ("Минимальный балл",        s["min_score"]),
        ("AI проверок",             s["ai_checks"]),
        ("Авто проверок",           s["auto_checks"]),
        ("Ручных проверок",         s["manual_checks"]),
        ("Провалено AI",            s["failed_ai"]),
    ]
    for row in rows:
        ws.append(row)
    _autowidth(ws)

    # ── Tests sheet ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Тесты")
    _header(ws2, ["Тест", "Вопросов", "Сессий", "Попыток", "Сред.балл", "Лучший балл", "% прохождения", "Прошли", "Провалили"])
    for t in data["tests_kpi"]:
        ws2.append([
            t["test_title"], t["questions"], t["sessions"],
            t["total_attempts"], t["avg_score"], t["best_score"],
            t["pass_rate"], t["passed"], t["failed"],
        ])
    _autowidth(ws2)

    # ── Students sheet ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Топ студентов")
    _header(ws3, ["Студент", "Попыток", "Сред.балл", "Лучший балл", "Последняя активность"])
    for st in data["top_students"]:
        ws3.append([
            st["student_name"], st["attempts"], st["avg_score"],
            st["best_score"], str(st["last_activity"] or ""),
        ])
    _autowidth(ws3)

    # ── AI sheet ──────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("AI Проверка")
    _header(ws4, ["Статус", "Количество"])
    ai = data["ai_kpi"]
    for k, v in [("Ожидают", ai["pending"]), ("Обрабатывается", ai["processing"]),
                 ("Готово", ai["done"]), ("Ошибка", ai["failed"])]:
        ws4.append([k, v])
    ws4.append(["Сред. AI оценка", ai["avg_ai_score"]])
    ws4.append(["Сред. уверенность AI", ai["avg_confidence"]])
    _autowidth(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _header(ws, cols: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)