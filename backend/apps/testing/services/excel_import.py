from __future__ import annotations
import io
from dataclasses import dataclass, field
import pandas as pd


def _get_models():
    from ..models import Test, Question, QuestionOption, QuestionType, DifficultyLevel
    return Test, Question, QuestionOption, QuestionType, DifficultyLevel


SHEET_NAME = 'questions'
REQUIRED_COLUMNS = {'test_title', 'question_text', 'type', 'difficulty'}
ALL_COLUMNS = [
    'test_title', 'question_text', 'type', 'language',
    'option_1', 'option_2', 'option_3', 'option_4',
    'correct', 'difficulty',
]
VALID_TYPES        = {'single_choice', 'multiple_choice', 'text', 'code'}
VALID_DIFFICULTIES = {'easy', 'medium', 'hard'}
VALID_LANGUAGES    = {'python', 'javascript', 'html', 'css', ''}

TEMPLATE_ROWS = [
    {
        'test_title': 'Python Basics', 'question_text': 'Что делает функция len()?',
        'type': 'single_choice', 'language': '',
        'option_1': 'Возвращает длину', 'option_2': 'Возвращает тип',
        'option_3': 'Возвращает значение', 'option_4': '',
        'correct': '1', 'difficulty': 'easy',
    },
    {
        'test_title': 'Python Basics', 'question_text': 'Какие типы данных изменяемы?',
        'type': 'multiple_choice', 'language': '',
        'option_1': 'list', 'option_2': 'tuple', 'option_3': 'dict', 'option_4': 'str',
        'correct': '1,3', 'difficulty': 'medium',
    },
    {
        'test_title': 'Python Basics',
        'question_text': 'Напишите функцию, возвращающую сумму двух чисел.',
        'type': 'code', 'language': 'python',
        'option_1': '', 'option_2': '', 'option_3': '', 'option_4': '',
        'correct': '', 'difficulty': 'hard',
    },
    {
        'test_title': 'Python Basics', 'question_text': 'Объясните, что такое декоратор.',
        'type': 'text', 'language': '',
        'option_1': '', 'option_2': '', 'option_3': '', 'option_4': '',
        'correct': '', 'difficulty': 'medium',
    },
]


# ── DTOs ──────────────────────────────────────────────────────────────────────

@dataclass
class RowError:
    row: int
    column: str
    message: str

    def __str__(self):
        return f'Строка {self.row}, «{self.column}»: {self.message}'


@dataclass
class ParsedQuestion:
    test_title:      str
    question_text:   str
    question_type:   str
    language:        str
    difficulty:      str
    options:         list
    correct_indices: list
    row_number:      int


@dataclass
class ImportPreview:
    rows:   list = field(default_factory=list)
    errors: list = field(default_factory=list)

    @property
    def is_valid(self):
        return len(self.errors) == 0

    @property
    def test_titles(self):
        return sorted({r.test_title for r in self.rows})

    @property
    def total_options(self):
        return sum(len(r.options) for r in self.rows)


@dataclass
class ImportResult:
    tests_created:     int = 0
    questions_created: int = 0
    questions_updated: int = 0
    options_created:   int = 0
    errors: list = field(default_factory=list)

    @property
    def success(self):
        return not self.errors


# ── Template ──────────────────────────────────────────────────────────────────

def generate_template() -> bytes:
    df = pd.DataFrame(TEMPLATE_ROWS, columns=ALL_COLUMNS)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        ws = writer.sheets[SHEET_NAME]
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='1F4E79')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or '')) for c in col) + 4, 40
            )
    output.seek(0)
    return output.read()


# ── Parse ─────────────────────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> ImportPreview:
    preview = ImportPreview()
    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes), sheet_name=SHEET_NAME, dtype=str, keep_default_na=False
        )
    except Exception as exc:
        preview.errors.append(RowError(0, 'файл', f'Не удалось прочитать файл: {exc}'))
        return preview

    df.columns = df.columns.str.strip().str.lower()
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        preview.errors.append(
            RowError(0, 'столбцы', f'Отсутствуют обязательные столбцы: {", ".join(sorted(missing))}')
        )
        return preview

    for col in ALL_COLUMNS:
        if col not in df.columns:
            df[col] = ''

    for idx, row in df.iterrows():
        _validate_row(row, idx + 2, preview)

    return preview


def _validate_row(row, excel_row: int, preview: ImportPreview):
    errors_before = len(preview.errors)

    def err(col, msg):
        preview.errors.append(RowError(excel_row, col, msg))

    test_title    = str(row.get('test_title', '')).strip()
    question_text = str(row.get('question_text', '')).strip()
    q_type        = str(row.get('type', '')).strip().lower()
    difficulty    = str(row.get('difficulty', '')).strip().lower()
    language      = str(row.get('language', '')).strip().lower()
    correct_raw   = str(row.get('correct', '')).strip()

    if not test_title:    err('test_title',    'Поле не может быть пустым.')
    if not question_text: err('question_text', 'Поле не может быть пустым.')

    if q_type not in VALID_TYPES:
        err('type', f'Допустимые значения: {", ".join(VALID_TYPES)}. Получено: «{q_type}».')
        return

    if difficulty not in VALID_DIFFICULTIES:
        err('difficulty', f'Допустимые значения: easy/medium/hard. Получено: «{difficulty}».')

    if language not in VALID_LANGUAGES:
        err('language', f'Допустимые значения: python, javascript, html, css или пусто. Получено: «{language}».')

    if q_type == 'code' and not language:
        err('language', 'Для вопроса типа «code» обязателен язык программирования.')
    if q_type != 'code' and language:
        err('language', 'Язык программирования указывается только для вопросов типа «code».')

    options    = [str(row.get(f'option_{i}', '')).strip() for i in range(1, 5)]
    non_empty  = [o for o in options if o]

    if q_type in ('single_choice', 'multiple_choice') and len(non_empty) < 2:
        err('option_1', 'Для вопроса с выбором ответа необходимо минимум 2 варианта.')

    correct_indices = []
    if q_type == 'single_choice':
        if not correct_raw:
            err('correct', 'Для single_choice укажите номер правильного ответа (1–4).')
        else:
            try:
                i = int(correct_raw) - 1
                if not (0 <= i <= 3):
                    raise ValueError
                correct_indices = [i]
            except ValueError:
                err('correct', f'Должно быть числом 1–4. Получено: «{correct_raw}».')

    elif q_type == 'multiple_choice':
        if not correct_raw:
            err('correct', 'Для multiple_choice укажите номера правильных ответов, например «1,3».')
        else:
            try:
                for p in correct_raw.split(','):
                    i = int(p.strip()) - 1
                    if not (0 <= i <= 3):
                        raise ValueError(f'Индекс {i + 1} вне диапазона 1–4.')
                    correct_indices.append(i)
            except ValueError as e:
                err('correct', f'Формат: «1,3». Ошибка: {e}')

    if len(preview.errors) == errors_before:
        preview.rows.append(ParsedQuestion(
            test_title=test_title,
            question_text=question_text,
            question_type=q_type,
            language=language,
            difficulty=difficulty,
            options=non_empty,
            correct_indices=correct_indices,
            row_number=excel_row,
        ))


# ── Commit (idempotent, no duplicates) ────────────────────────────────────────

def commit_import(preview: ImportPreview) -> ImportResult:
    """
    Идемпотентный импорт:
    - Test:           get_or_create по title
    - Question:       update_or_create по (test, text)
    - QuestionOption: пересоздаются для каждого вопроса (delete + bulk_create)
    Всё в одной транзакции.
    """
    from django.db import transaction
    Test, Question, QuestionOption, _, __ = _get_models()

    result = ImportResult()

    try:
        with transaction.atomic():
            test_cache: dict = {}

            for pq in preview.rows:
                # ── Test ──────────────────────────────────────────────────────
                if pq.test_title not in test_cache:
                    test, created = Test.objects.get_or_create(
                        title=pq.test_title,
                        defaults={'level': pq.difficulty},
                    )
                    if created:
                        result.tests_created += 1
                    test_cache[pq.test_title] = test

                test = test_cache[pq.test_title]

                # ── Question (update_or_create by test+text) ──────────────────
                question, q_created = Question.objects.update_or_create(
                    test=test,
                    text=pq.question_text,
                    defaults={
                        'question_type': pq.question_type,
                        'language':      pq.language,
                        'difficulty':    pq.difficulty,
                    },
                )
                if q_created:
                    result.questions_created += 1
                else:
                    result.questions_updated += 1

                # ── Options: пересоздаём только для choice-вопросов ───────────
                if pq.question_type in ('single_choice', 'multiple_choice'):
                    question.options.all().delete()
                    new_opts = [
                        QuestionOption(
                            question=question,
                            text=text,
                            is_correct=(order - 1) in pq.correct_indices,
                            order=order,
                        )
                        for order, text in enumerate(pq.options, start=1)
                    ]
                    QuestionOption.objects.bulk_create(new_opts)
                    result.options_created += len(new_opts)

    except Exception as exc:
        result.errors.append(str(exc))

    return result


# ── Export ────────────────────────────────────────────────────────────────────

def export_questions_to_excel(test_ids=None) -> bytes:
    _, Question, __, ___, ____ = _get_models()

    qs = Question.objects.select_related('test').prefetch_related('options').order_by('test__title', 'order')
    if test_ids:
        qs = qs.filter(test_id__in=test_ids)

    rows = []
    for q in qs:
        opts        = list(q.options.order_by('order'))
        correct_nums = [str(i + 1) for i, o in enumerate(opts) if o.is_correct]
        rows.append({
            'test_title':    q.test.title,
            'question_text': q.text,
            'type':          q.question_type,
            'language':      q.language,
            'option_1':      opts[0].text if len(opts) > 0 else '',
            'option_2':      opts[1].text if len(opts) > 1 else '',
            'option_3':      opts[2].text if len(opts) > 2 else '',
            'option_4':      opts[3].text if len(opts) > 3 else '',
            'correct':       ','.join(correct_nums),
            'difficulty':    q.difficulty,
        })

    df = pd.DataFrame(rows, columns=ALL_COLUMNS) if rows else pd.DataFrame(columns=ALL_COLUMNS)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        ws = writer.sheets[SHEET_NAME]
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='1F4E79')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or '')) for c in col) + 4, 50
            )
    output.seek(0)
    return output.read()